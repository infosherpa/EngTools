from openpyxl import Workbook
from django.http import HttpResponse
import os
from django.conf import settings
import re
from ..models import LoadDefinition
from decimal import Decimal


def create_workbook(tunnel_frame, auth=False):
    """Create excel workbook of relevant data for import into SAP2000"""
    tunnel_frame.get_frame_geometry()
    wb = Workbook()
    loads = LoadDefinition.objects.filter(parent_frame=tunnel_frame.id)
    # Table definition for SAP2000 Compatibility and system variables ie. Dimension system (mm, m, in)

    dest_filename = f"{tunnel_frame.frame_description}.xlsx"

    # Active Degrees of Freedom

    ws1 = wb.create_sheet(title="ACTIVE DEGREES OF FREEDOM")
    ws1['A1'] = "TABLE: ACTIVE DEGREES OF FREEDOM"
    degrees_freedom_headers = [
        ["UX", "UY", "UZ", "RX", "RY", "RZ"],
        ["Yes/No", "Yes/No", "Yes/No", "Yes/No", "Yes/No", "Yes/No"],
        ["Yes", "Yes", "Yes", "Yes", "Yes", "Yes"],
    ]
    for row in degrees_freedom_headers:
        ws1.append(row)

    # Analysis Options
    ########

    ws2 = wb.create_sheet(title="ANALYSIS OPTIONS")
    ws2['A1'] = "TABLE: ANALYSIS OPTIONS"
    analysis_headers = [
        ["Solver", "SolverProc", "Force32Bit", "StiffCase", "GeomMod", "HingeOpt", "NumAThreads", "MaxFileSize", "NumDThreads", "NumRThreads", "UseMMFiles", "AllowDiff"],
        ["Text", "Text", "Yes/No", "Text", "Yes/No", "Text", "Unitless", "Unitless", "Unitless", "Unitless", "Text", "Yes/No"],
        ["Advanced", "Auto", "No", "", "", "In Elements", 0, 0, 0, 0, "Program Determined", "No"]
    ]
    for h in analysis_headers:
        ws2.append(h)

    #### Case 1 - Load Assignments
    ####### Static Load Case

    """ws3 = wb.create_sheet(title="Case - Static 1 - Load Assigns")
    ws3['A1'] = "Table: Case - Static 1 - Load Assignments"
    ws3.append(['Case', 'LoadType', 'LoadName', 'LoadSF'])
    ws3.append(['Text', 'Text', 'Text', 'Unitless'])

    
    for load in loads:
        ws3.append([load.load_pattern_description, 'Load pattern', load.load_pattern_description, 1])"""

    ##
    # WS4 Frame Connectivity (Members)

    ws4 = wb.create_sheet(title="Connectivity - Frame")
    ws4['A1'] = "TABLE: Connectivity - Frame"
    connectivity_frame_headers = [
        ["Frame", "JointI", "JointJ", "IsCurved", "Length", "CentroidX", "CentroidY", "CentroidZ", "GUID"],
        ["Text", "Text", "Text", "Yes/No", "mm", "mm", "mm", "mm", "Text"]
    ]
    for values in connectivity_frame_headers:
        ws4.append(values)
    for memb_key, member in tunnel_frame.connectivity_frame.items():
        ws4.append([memb_key, member[0], member[1], "No"])

    #### Co Ordinates System

    ws5 = wb.active
    ws5.title = "Coordinate Systems"
    ws5['A1'] = "TABLE:  Coordinate Systems"
    co_ordinate_headings = [
        ["Name", "Type", "X", "Y", "Z", "AboutZ", "AboutY", "AboutX"],
        ["Text", "Text", "mm", "mm", "mm", "Degrees", "Degrees", "Degrees"],
        ["GLOBAL", "Cartesian", 0, 0, 0, 0, 0, 0]
    ]
    for row in co_ordinate_headings:
        ws5.append(row)

    ########

    ws6 = wb.create_sheet(title="Frame Auto Mesh")
    ws6['A1'] = "TABLE: FRAME AUTO MESH ASSIGNMENTS"
    ws6.append(["Frame", "AutoMesh", "AtJoints", "AtFrames", "NumSegments", "MaxLength", "MaxDegrees"])
    ws6.append(["Text", "Yes/No", "Yes/No", "Yes/No", "Unitless", "mm", "Degrees"])
    for frm_num, frame in tunnel_frame.connectivity_frame.items():
        ws6.append([frm_num, "Yes", "Yes", "No", 0, 500, 0])

    ######
    # Frame Design Procedures

    ws7 = wb.create_sheet(title="FRAME DESIGN PROCEDURES")
    ws7['A1'] = "TABLE: FRAME DESIGN PROCEDURES"
    ws7.append(["Frame", "DesignProc"])
    ws7.append(["Text", "Text"])
    for frm_num, frame in tunnel_frame.connectivity_frame.items():
        ws7.append([frm_num, "From Material"])

    #######

    ws8 = wb.create_sheet(title="FRAME LOAD TRANSFER OPTIONS")
    ws8['A1'] = "TABLE: FRAME LOAD TRANSFER OPTIONS"
    ws8.append(["Frame", "Transfer"])
    ws8.append(["Text", "Yes/No"])
    for frm_num, frame in tunnel_frame.connectivity_frame.items():
        ws8.append([frm_num, "Yes"])

    ######
    # Frame Loads

    ws9 = wb.create_sheet(title="Frame Loads - Distributed")
    ws9['A1'] = "Table: Frame Loads - Distributed"
    ws9.append(['Frame', 'LoadPat', 'CoordSys', 'Type', 'Dir', 'DistType', 'RelDistA', 'RelDistB', 'AbsDistA', 'AbsDistB', 'FOverLA', 'FOverLB', 'GUID'])
    ws9.append(['Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Unitless', 'Unitless', 'mm', 'mm', 'KN/m', 'KN/m', 'text'])

    for load in loads:
        if load.load_location == "RS":
            x = {k: v for (k, v) in tunnel_frame.connectivity_frame.items() if v[2][0:2] == "RS"}
            for key, value in x.items():
                ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0, 1, 0,
                             tunnel_frame.joint_coordinates[value[1]][0] - tunnel_frame.joint_coordinates[value[0]][0],load.start_force, load.end_force, ""])
        elif load.load_location == "CS":
            x = {k: v for (k, v) in tunnel_frame.connectivity_frame.items() if v[2][0:2] == "CS" and (tunnel_frame.frame_outer_width - tunnel_frame.wall_slab_thickness) >= tunnel_frame.joint_coordinates[v[1]][0] > tunnel_frame.wall_slab_thickness}
            for key, value in x.items():
                ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0, 1, 0,
                         tunnel_frame.joint_coordinates[value[1]][0] - tunnel_frame.joint_coordinates[value[0]][0],
                         load.start_force, load.end_force, ""])
        elif load.load_location == "LW":
            x = {k: v for (k, v) in tunnel_frame.connectivity_frame.items() if v[2][0:2] == "WS" and tunnel_frame.joint_coordinates[v[0]][0]<tunnel_frame.frame_inner_width}
            outer_length = tunnel_frame.frame_outer_height

            ###
            # If we have a starting depth for force action we must have a routine to match the force to the geometry of the frame.

            if load.force_start_depth is not None:
                active_force = 0
                active_depth = Decimal(0)
                force_action_length = Decimal(load.force_end_depth - load.force_start_depth)
                force_change = Decimal(load.end_force - load.start_force)

                for key, value in x.items():
                    member_length = tunnel_frame.joint_coordinates[value[0]][1] - \
                                    tunnel_frame.joint_coordinates[value[1]][1]

                    if load.force_start_depth >= (
                            tunnel_frame.frame_outer_height - tunnel_frame.joint_coordinates[value[1]][1]):
                        continue
                    else:
                        if active_depth == 0:
                            starting_abs_length = Decimal(load.force_start_depth) - (
                                        tunnel_frame.frame_outer_height - tunnel_frame.joint_coordinates[value[0]][1])
                            starting_rel_length = starting_abs_length / member_length
                            starting_abs_force = Decimal(load.start_force)
                        else:
                            starting_abs_length = 0
                            starting_abs_force = active_force
                            starting_rel_length = 0
                    # print(tunnel_frame.frame_outer_height - load.force_end_depth)
                    # print(tunnel_frame.joint_coordinates[value[1]][1])
                    if tunnel_frame.frame_outer_height - load.force_end_depth < \
                            tunnel_frame.joint_coordinates[value[1]][1]:

                        final_rel_len = 1
                        final_abs_len = abs(
                            tunnel_frame.joint_coordinates[value[1]][1] - tunnel_frame.joint_coordinates[value[0]][1])
                        final_abs_force = Decimal(load.start_force) + (active_depth + (
                                    final_abs_len - starting_abs_length)) / force_action_length * force_change

                        active_depth = active_depth + (final_abs_len - starting_abs_length)
                        active_force = final_abs_force
                        ws9.append(
                            [key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist",
                             starting_rel_length, final_rel_len,
                             starting_abs_length, final_abs_len, starting_abs_force, final_abs_force, ""])

                    else:
                        final_rel_len = abs((tunnel_frame.frame_outer_height - load.force_end_depth) - (
                        tunnel_frame.joint_coordinates[value[0]][1])) / member_length
                        final_abs_force = load.end_force
                        final_abs_len = final_rel_len * member_length
                        ws9.append(
                            [key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist",
                             starting_rel_length, final_rel_len,
                             starting_abs_length, final_abs_len, starting_abs_force, final_abs_force, ""])
                        break

            else:

                length_100 = 0
                length_25 = 0
                length_75 = 0
                length_0 = 0

                for key, value in x.items():
                    length_0 = 0
                    length_25 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])*Decimal(0.25)
                    length_75 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])*Decimal(0.75)
                    length_100 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])
                    force_0 = (outer_length-tunnel_frame.joint_coordinates[value[0]][1])/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_25 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+Decimal(0.25)*(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_75 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+Decimal(0.75)*(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_100 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0, 0.25, length_0, length_25, force_0, force_25, ""])
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0.25, 0.75, length_25, length_75, force_25, force_75, ""]),
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0.75, 1, length_75, length_100, force_75, force_100, ""])

        elif load.load_location == "RW":
            length_100 = 0
            length_25 = 0
            length_75 = 0
            length_0 = 0
            x = {k: v for (k, v) in tunnel_frame.connectivity_frame.items() if v[2][0:2] == "WS" and tunnel_frame.joint_coordinates[v[0]][0]>tunnel_frame.frame_inner_width}
            outer_length = tunnel_frame.frame_outer_height

            if load.force_start_depth is not None:
                active_force = 0
                active_depth = Decimal(0)
                force_action_length = Decimal(load.force_end_depth-load.force_start_depth)
                force_change = Decimal(load.end_force - load.start_force)

                for key, value in x.items():
                    member_length = tunnel_frame.joint_coordinates[value[0]][1] - tunnel_frame.joint_coordinates[value[1]][1]

                    if load.force_start_depth >= (tunnel_frame.frame_outer_height - tunnel_frame.joint_coordinates[value[1]][1]):
                        continue
                    else:
                        if active_depth == 0:
                            starting_abs_length = Decimal(load.force_start_depth) - (tunnel_frame.frame_outer_height - tunnel_frame.joint_coordinates[value[0]][1])
                            starting_rel_length = starting_abs_length/member_length
                            starting_abs_force = Decimal(load.start_force)
                        else:
                            starting_abs_length = 0
                            starting_abs_force = active_force
                            starting_rel_length = 0
                    # print(tunnel_frame.frame_outer_height - load.force_end_depth)
                    # print(tunnel_frame.joint_coordinates[value[1]][1])
                    if tunnel_frame.frame_outer_height - load.force_end_depth < tunnel_frame.joint_coordinates[value[1]][1]:

                        final_rel_len = 1
                        final_abs_len = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])
                        final_abs_force = Decimal(load.start_force)+(active_depth+(final_abs_len-starting_abs_length))/force_action_length*force_change

                        active_depth = active_depth + (final_abs_len-starting_abs_length)
                        active_force = final_abs_force
                        ws9.append(
                            [key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist",
                             starting_rel_length, final_rel_len,
                             starting_abs_length, final_abs_len, -starting_abs_force, -final_abs_force, ""])

                    else:
                        final_rel_len = abs((tunnel_frame.frame_outer_height - load.force_end_depth)-(tunnel_frame.joint_coordinates[value[0]][1]))/member_length
                        final_abs_force = load.end_force
                        final_abs_len = final_rel_len*member_length
                        ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", starting_rel_length, final_rel_len,
                                starting_abs_length, final_abs_len, -starting_abs_force, -final_abs_force, ""])
                        break

            else:
                for key, value in x.items():
                    print(tunnel_frame.joint_coordinates[value[0]][0])
                    length_0 = 0
                    length_25 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])*Decimal(0.25)
                    length_75 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])*Decimal(0.75)
                    length_100 = abs(tunnel_frame.joint_coordinates[value[1]][1]-tunnel_frame.joint_coordinates[value[0]][1])
                    force_0 = (outer_length-tunnel_frame.joint_coordinates[value[0]][1])/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_25 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+Decimal(0.25)*(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_75 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+Decimal(0.75)*(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    force_100 = (tunnel_frame.frame_outer_height-tunnel_frame.joint_coordinates[value[0]][1]+(tunnel_frame.joint_coordinates[value[0]][1]-tunnel_frame.joint_coordinates[value[1]][1]))/tunnel_frame.frame_outer_height*Decimal(load.end_force)
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0, 0.25, length_0, length_25, -force_0, -force_25, ""])
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0.25, 0.75, length_25, length_75, -force_25, -force_75, ""]),
                    ws9.append([key, load.load_pattern_description, "GLOBAL", "Force", load.load_direction, "RelDist", 0.75, 1, length_75, length_100, -force_75, -force_100, ""])

    ####### Frame Section Properties

    ws10 = wb.create_sheet(title="Frame Props 01 - General")
    ws10['A1'] = "TABLE: FRAME SECTION PROPERTIES 01 - GENERAL"
    headers = ["SectionName", "Material", "Shape", "t3", "t2", "Area", "TorsConst", "I33", "I22", "I23", "AS2", "AS3", "S33", "S22", "Z33", "Z22", "R33", "R22", "Color", "FromFile", "AMod", "A2Mod", "A3Mod", "JMod", "I2Mod", "I3Mod", "MMod", "WMod", "GUID", "Notes"]
    ws10.append(headers)
    ws10.append(["Text", "Text", "Text", "mm", "mm", "mm2", "mm4", "mm4", "mm4", "mm4", "mm2", "mm2", "mm3", "mm3", "mm3", "mm3", "mm", "mm", "Text", "Yes/No", "Unitless", "Unitless", "Unitless", "Unitless", "Unitless", "Unitless", "Unitless", "Unitless", "Text", "Text"])


    ######## Frame Section Assignment

    ws11 = wb.create_sheet(title="Frame Section Assignments")
    ws11['A1'] = "TABLE: FRAME SECTION ASSIGNMENTS"
    headers = ["Frame", "SectionType", "AutoSelect", "AnalSect", "DesignSect", "MatProp"]
    ws11.append(headers)
    ws11.append(["Text", "Text", "Text", "Text", "Text", "Text"])

    members_name = []

    ### Routine for Frame section and assignment

    for frm_num, member in tunnel_frame.connectivity_frame.items():
        try:
            thickness = re.findall(r"[\d]+", member[2])[0]
        except:
            thickness = 1
        slabs = ("RS", "IS", "CS")

        if member[2][:2] in slabs:
            strength = tunnel_frame.concrete_strength_slabs
            modifier = tunnel_frame.slab_stiffness_modifier
        elif member[2][:2] == "WS":
            strength = tunnel_frame.concrete_strength_walls
            modifier = tunnel_frame.wall_stiffness_modifier
        elif member[2][:3] == "COL":
            strength = tunnel_frame.concrete_strength_columns
            modifier = tunnel_frame.column_stiffness_modifier
        else:
            break
        if member[2][-1]=="S":
            modifier = 1
        vals = [member[2], f"{strength}Psi", "Rectangular", thickness, 1000, "", "", "", "", "", "", "", "", "", "", "",
                "", "", "", "No", 1, 1, 1, modifier, modifier, modifier, 1, 1, "", ""]
        ws10.append(vals)

        # Frame Section Properties routine

        vals = [frm_num, "Rectangular", "N.A.", member[2], member[2], "Default"]
        ws11.append(vals)

    ####### Frame Spring Assignment

    ws12 = wb.create_sheet(title="Frame Spring Assignments")
    ws12['A1'] = "Table: Frame Spring Assignments"
    ws12.append(['Frame', 'Type', 'Stiffness', 'SimpleType', 'Dir1Type', 'CoordSys', 'VecX', 'VecY', 'VecZ'])
    ws12.append(['Text', 'Text', 'kN/m/m', 'Text', 'Text', 'Text', 'Unitless', 'Unitless', 'Unitless'])
    vecy = 0
    for frm_num, member in tunnel_frame.connectivity_frame.items():
        if member[2][0:2]=="RS":
            continue
        elif member[2][0:2]=="CS":
            continue
        elif member[2][0:2]=="IS":
            stiffness = 55000
            vecx = 0
            vecz = 1
        elif member[2][0:2]=="WS":
            stiffness = 20000
            vecz = 0
            if tunnel_frame.joint_coordinates[member[0]][0] < tunnel_frame.frame_inner_width:
                vecx = 1
            elif tunnel_frame.joint_coordinates[member[0]][0] > tunnel_frame.frame_inner_width:
                vecx = -1
        elif member[2][0:2]=="CO":
            continue
        ws12.append([frm_num, 'Simple', stiffness, 'Compression Only', 'User Vector', 'GLOBAL', vecx, vecy, vecz])

    #######

    ws13 = wb.create_sheet(title="GRID LINES")
    ws13['A1'] = "TABLE: GRID LINES"
    ws13.append(["CoordSys", "AxisDir", "GridID", "XRYZCoord", "LineType", "LineColor", "Visible", "BubbleLoc", "AllVisible", "BubbleSize"])
    ws13.append(
        ["Text", "Text", "Text", "mm", "Text", "Text", "Yes/No", "Text", "Yes/No", "mm"])
    i = 65
    for grid in tunnel_frame.grid_locations_x:
        ws13.append(["GLOBAL", "X", chr(i), float(grid), "Primary", "Gray8Dark", "Yes", "End", "Yes", 1250])
        i += 1
    ws13.append(["GLOBAL", "Y", 1, 0, "Primary", "Gray8Dark", "Yes", "End", "Yes", 1250])

    i = 1
    for grid in tunnel_frame.grid_locations_z:
        ws13.append(["GLOBAL", "Z", f"Z{i}", float(grid), "Primary", "Gray8Dark", "Yes", "Start", "Yes", 1250])
        i+=1

    # Joint Co-Ordinates

    ws14 = wb.create_sheet(title="Joint Coordinates")
    ws14['A1'] = "TABLE:  Joint Coordinates"
    joint_coord_headings = [
        ["Joint", "CoordSys", "CoordType", "XorR", "Y", "Z", "SpecialJT", "GlobalX", "GlobalY", "GlobalZ", "GUID"],
        ["Text", "Text", "Text", "mm", "mm", "mm", "Yes/No", "mm", "mm", "mm", "Text"]
    ]
    for row in joint_coord_headings:
        ws14.append(row)

    vertexes = tunnel_frame.joint_coordinates
    for vrtx_num, values in vertexes.items():
        ws14.append(
            [vrtx_num, "GLOBAL", "Cartesian", values[0], 0, values[1], "No", values[0], 0, values[1]]
        )

    #########

    ws15 = wb.create_sheet(title="JOINT PATTERN DEFINITIONS")
    ws15['A1'] = "TABLE: JOINT PATTERN DEFINITIONS"
    ws15.append(["Pattern"])
    ws15.append(["Text"])
    ws15.append(["Default"])

    # Load Case Definitions

    ws16 = wb.create_sheet(title="Load Case Definitions")
    ws16['A1'] = "Table: Load Case Definitions"
    ws16.append(
        ['Case', 'Type', 'InitialCond', 'ModalCase', 'BaseCase', 'MassSource', 'DesTypeOpt', 'DesignType', 'DesActOpt', 'DesignAct', 'AutoType', 'RunCase', 'CaseStatus', 'GUID', 'Notes'])
    ws16.append(['Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Yes/No', 'Text', 'Text', 'Text'])
    for load in loads:
        if load.load_pattern_description[0:2] == "LL":
            load_type = "Live"
            load_type2 = "Short-Term Composite"
        else:
            load_type = "Other"
        ws16.append([load.load_pattern_description, "NonStatic", "Zero", "", "", "", "Prog Det", load_type, "Prog Det", "", "None", "Yes", "Not Run", "", ""])

    ######

    ws17 = wb.create_sheet(title="Load Pattern Definitions")
    ws17['A1'] = "Table: Load Pattern Definitions"
    ws17.append(
        ['LoadPat', 'DesignType', 'SelfWtMult', 'AutoLoad', 'GUID', 'Notes'])
    ws17.append(['Text', 'Text', 'Unitless', 'Text', 'Text', 'Text'])
    ws17.append(["DEAD", "Dead", 1, "", "", ""])
    loads_list = []
    for load in loads:
        if load.load_pattern_description in loads_list:
            continue
        if load.load_pattern_description == "LL_RS":
            design_type = "Live"
        elif load.load_pattern_description == "H_L" or load.load_pattern_description == "R_L":
            design_type = "Other"
        ws17.append([load.load_pattern_description, design_type, 0, "", "", ""])
        loads_list.append(load.load_pattern_description)

    # WS3 - Program Control SAP2000 required table

    ws18 = wb.create_sheet(title="Program Control")
    ws18['A1'] = "TABLE: Program Control"
    program_control_values = [
        ["ProgramName", "Version", "ProgLevel", "LicenseNum", "LicenseOS", "LicenseSC", "LicenseHT", "CurrUnits",
         "SteelCode", "ConcCode", "AlumCode", "ColdCode", "RegenHinge"],
        ["Text", "Text", "Text", "Text", "Yes/No", "Yes/No", "Yes/No", "Text", "Text", "Text", "Text", "Text", "Yes/No"],
        ["SAP2000", "23.2.0", "Ultimate", "", "Yes", "Yes", "No", "N, mm, C", "AISC 360-16", "ACI 318-14", "AA 2015",
         "AISI-16", "Yes"]
    ]
    for row in program_control_values:
        ws18.append(row)


    #####

    ws19 = wb.create_sheet(title="MatProp 01 - General")
    ws19['A1'] = "TABLE: Material Properties 01 - General"
    material_properties_headers = [
        ["Material", "Type", "Grade", "SymType", "TempDepend", "Color", "GUID", "Notes"],
        ["Text", "Text", "Text", "Text", "Yes/No", "Text", "Text", "Text"],
    ]
    for row in material_properties_headers:
        ws19.append(row)

    ws19.append([f"{tunnel_frame.concrete_strength_walls}Psi", "Concrete", f"f'c {tunnel_frame.concrete_strength_walls} psi",
                     "Isotropic", "No", "Yellow", "", ""])
    if tunnel_frame.concrete_strength_walls != tunnel_frame.concrete_strength_slabs:
        ws19.append([f"{tunnel_frame.concrete_strength_slabs}Psi", "Concrete", f"f'c {tunnel_frame.concrete_strength_slabs} psi",
                     "Isotropic", "No", "Yellow", "", ""])
    if tunnel_frame.column_width and tunnel_frame.concrete_strength_columns != tunnel_frame.concrete_strength_slabs:
        ws19.append([f"{tunnel_frame.concrete_strength_columns}Psi", "Concrete",
                     f"f'c {tunnel_frame.concrete_strength_columns} psi",
                     "Isotropic", "No", "Yellow", "", ""])

    ws20 = wb.create_sheet(title="Case - Static 1 - Load Assigns")
    ws20['A1'] = "Case - Static 1 - Load Assignments"
    ws20.append(['Case', 'LoadType', 'LoadName', 'LoadSF'])
    ws20.append(['Text', 'Text', 'Text', 'Unitless'])

    #####

    """ws21 = wb.create_sheet(title="Joint Restraint Assignments")
    ws21['A1'] = "TABLE: Joint Restraint Assignments"
    ws21.append(['Joint', 'U1', 'U2', 'U3', 'R1', 'R2', 'R3'])
    ws21.append(['Text', 'Yes/No', 'Yes/No', 'Yes/No', 'Yes/No', 'Yes/No', 'Yes/No'])
    for joint, values in tunnel_frame.joint_coordinates.items():
        if values[1] == 0:
            ws21.append([joint, 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'])"""

    loads_list = []
    for load in loads:
        if load.load_pattern_description in loads_list:
            continue
        ws20.append([load.load_pattern_description, 'Load pattern', load.load_pattern_description, 1])
        loads_list.append(load.load_pattern_description)


    if auth is False:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename={dest_filename}'
        wb.save(response)
        return response

    target = os.path.join(settings.STATIC_DIR, "XLXS")
    wb.save(target+f"/{dest_filename}")
    wb_location = os.path.join(target, dest_filename)

    return wb_location